import openpyxl

wb = openpyxl.load_workbook(r"C:\Deepseek\Test-1\Personal Técnico 2026.xlsx", data_only=True)
print("Hojas:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n=== Hoja: {ws.title}  (filas {ws.max_row}, cols {ws.max_column}) ===")
    rows = list(ws.iter_rows(values_only=True))
    # Cabecera (primeras filas no vacías)
    for i, r in enumerate(rows[:12]):
        print(i, [str(c)[:30] if c is not None else "" for c in r])
